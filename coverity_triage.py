#!/usr/bin/env python3
"""
Coverity Triage GUI — Analyse an HTML report folder, review/edit dispositions,
and export to Excel.

Usage:
    python coverity_gui_excel.py
"""

import csv
import hashlib
import json
import logging
import os
import re
import sys
import threading
import tkinter as tk
from datetime import datetime, timezone
from tkinter import filedialog, messagebox, scrolledtext, ttk

# ---------------------------------------------------------------------------
# Optional: make sure local modules are importable when launched from any cwd
# ---------------------------------------------------------------------------
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)

from context_builder import build_defect_context, warm_workspace_index
from heuristic_analyzer import analyze_defect
from html_report_parser import parse_coverity_html
from checker_categories import category_for_checker, CATEGORY_ORDER

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Column definitions for the main table
# ---------------------------------------------------------------------------
COLUMNS = ("CID", "Checker", "Category", "File", "Line", "Disposition", "Confidence", "Comment")
COL_WIDTHS = {"CID": 60, "Checker": 160, "Category": 140, "File": 260, "Line": 55,
              "Disposition": 130, "Confidence": 70, "Comment": 300}

DISPOSITION_OPTIONS = ["Bug", "False positive", "Intentional", "Needs review"]

DISP_COLORS = {
    "Bug":           "#ffd6d6",
    "False positive":"#d6ffd6",
    "Intentional":   "#d6e8ff",
    "Needs review":  "#fff9d6",
}


def _iid_safe(text):
    """Return a Treeview-safe iid string derived from ``text``."""
    return re.sub(r"[^A-Za-z0-9_.-]", "_", str(text)) or "item"


# ===========================================================================
class CoverityExcelApp:
    """Main application window."""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Coverity Triage — Excel Export")
        self.root.geometry("1150x680")
        self.root.resizable(True, True)

        self._defect_rows: list[dict] = []   # processed results
        self._raw_defects: list[dict] = []   # parsed defects before analysis
        self._cat_items: dict[str, str] = {}  # category name -> tree parent iid

        self._build_toolbar()
        self._build_table()
        self._build_statusbar()

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------
    def _build_toolbar(self):
        tb = ttk.Frame(self.root, padding=6)
        tb.pack(fill="x")

        # Report folder
        ttk.Label(tb, text="HTML Report Folder:").grid(row=0, column=0, sticky="w")
        self.report_var = tk.StringVar()
        ttk.Entry(tb, textvariable=self.report_var, width=45).grid(row=0, column=1, padx=4)
        ttk.Button(tb, text="Browse…", command=self._browse_report).grid(row=0, column=2)

        # Source root
        ttk.Label(tb, text="Source Root (optional):").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.src_root_var = tk.StringVar()
        ttk.Entry(tb, textvariable=self.src_root_var, width=45).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Button(tb, text="Browse…", command=self._browse_src).grid(row=1, column=2, pady=(4, 0))

        # Language
        ttk.Label(tb, text="Language:").grid(row=0, column=3, padx=(20, 4), sticky="w")
        self.lang_var = tk.StringVar(value="c")
        ttk.Combobox(tb, textvariable=self.lang_var, values=["c", "cpp"],
                     width=6, state="readonly").grid(row=0, column=4, sticky="w")

        # Limit
        ttk.Label(tb, text="Limit (0=all):").grid(row=1, column=3, padx=(20, 4), sticky="w")
        self.limit_var = tk.StringVar(value="0")
        ttk.Entry(tb, textvariable=self.limit_var, width=6).grid(row=1, column=4, sticky="w")

        # Action buttons
        btn_frame = ttk.Frame(tb)
        btn_frame.grid(row=0, column=5, rowspan=2, padx=(30, 0))

        self.run_btn = ttk.Button(btn_frame, text="▶  Run Analysis", command=self._run_analysis)
        self.run_btn.pack(fill="x", pady=2)

        ttk.Button(btn_frame, text="💾  Export to Excel", command=self._export_excel).pack(fill="x", pady=2)

        # Progress bar
        self.progress = ttk.Progressbar(self.root, mode="determinate")
        self.progress.pack(fill="x", padx=6, pady=(0, 2))

    def _build_table(self):
        frame = ttk.Frame(self.root)
        frame.pack(fill="both", expand=True, padx=6)

        self.tree = ttk.Treeview(frame, columns=COLUMNS, show="headings",
                                 selectmode="browse")
        for col in COLUMNS:
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._sort_by(c))
            self.tree.column(col, width=COL_WIDTHS[col], anchor="w",
                             stretch=(col == "Comment"))

        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        # Configure row tag colours
        for disp, colour in DISP_COLORS.items():
            self.tree.tag_configure(disp, background=colour)
        self.tree.tag_configure("cat_header", font=("Segoe UI", 9, "bold"))

        self.tree.bind("<Double-1>", self._on_row_double_click)

    def _build_statusbar(self):
        self.status_var = tk.StringVar(value="Ready. Browse to a Coverity HTML report folder and click Run.")
        ttk.Label(self.root, textvariable=self.status_var,
                  anchor="w", relief="sunken").pack(fill="x", padx=6, pady=(2, 4))

    # ------------------------------------------------------------------
    # Browse helpers
    # ------------------------------------------------------------------
    def _browse_report(self):
        folder = filedialog.askdirectory(title="Select Coverity HTML Report Folder")
        if folder:
            self.report_var.set(folder)

    def _browse_src(self):
        folder = filedialog.askdirectory(title="Select Source Code Root Folder")
        if folder:
            self.src_root_var.set(folder)

    # ------------------------------------------------------------------
    # Analysis
    # ------------------------------------------------------------------
    def _run_analysis(self):
        report_path = self.report_var.get().strip()
        if not report_path:
            messagebox.showwarning("Missing Input", "Please select a Coverity HTML report folder.")
            return
        if not os.path.exists(report_path):
            messagebox.showerror("Not Found", f"Path does not exist:\n{report_path}")
            return

        src_root = self.src_root_var.get().strip() or report_path
        language = self.lang_var.get()
        try:
            limit = int(self.limit_var.get())
        except ValueError:
            limit = 0

        self.run_btn.configure(state="disabled")
        self._clear_table()
        self._set_status("Parsing HTML report…")

        thread = threading.Thread(
            target=self._analysis_worker,
            args=(report_path, src_root, language, limit),
            daemon=True,
        )
        thread.start()

    def _analysis_worker(self, report_path, src_root, language, limit):
        try:
            defects = parse_coverity_html(report_path)
        except Exception as exc:
            self.root.after(0, lambda exc=exc: messagebox.showerror("Parse Error", str(exc)))
            self.root.after(0, lambda: self.run_btn.configure(state="normal"))
            return

        total = len(defects) if not limit else min(len(defects), limit)
        self.root.after(0, lambda: self._set_status(f"Analysing {total} defects…"))
        self.root.after(0, lambda: self.progress.configure(maximum=max(total, 1), value=0))

        # Warm the one-time workspace index up front so the first defect does not
        # silently stall the run (with no progress).
        if src_root and os.path.isdir(src_root):
            self.root.after(0, lambda: self._set_status("Indexing source tree once (cached)…"))
            try:
                warm_workspace_index(src_root, language)
            except Exception:
                pass
            self.root.after(0, lambda: self._set_status(f"Analysing {total} defects…"))

        results = []
        for i, defect in enumerate(defects):
            if limit and i >= limit:
                break
            cid = defect["cid"]
            checker = defect["checker"]
            first_event = defect["events"][0] if defect["events"] else {}
            file_path = first_event.get("file", "")
            line = first_event.get("line", "")

            try:
                context = build_defect_context(defect, src_root, language)
                if not context["function_code"]:
                    classification, comment = "Needs review", "Context extraction failed"
                else:
                    classification, comment, fix, confidence = analyze_defect(
                        context, checker, defect["events"]
                    )
            except Exception as exc:
                classification, comment = "Needs review", f"Error: {exc}"

            row = {
                "CID": cid,
                "Checker": checker,
                "Category": category_for_checker(checker),
                "File": file_path,
                "Line": line,
                "Disposition": classification,
                "Confidence": f"{int(confidence*100)}%",
                "Comment": comment,
                "_events": defect["events"],
            }
            results.append(row)

            # update UI from main thread
            idx = i
            row_copy = dict(row)
            self.root.after(0, lambda r=row_copy, i=idx, n=idx + 1: self._append_row(r, i, n))

        self._defect_rows = results
        self.root.after(0, self._analysis_done)

    def _analysis_done(self):
        count = len(self._defect_rows)
        bugs = sum(1 for r in self._defect_rows if r["Disposition"] == "Bug")
        fp = sum(1 for r in self._defect_rows if r["Disposition"] == "False positive")
        intentional = sum(1 for r in self._defect_rows if r["Disposition"] == "Intentional")
        needs = sum(1 for r in self._defect_rows if r["Disposition"] == "Needs review")
        self._set_status(
            f"Done — {count} defects  |  Bug: {bugs}  |  False positive: {fp}  "
            f"|  Intentional: {intentional}  |  Needs review: {needs}  "
            f"  (double-click a row to edit)"
        )
        self.progress.configure(value=self.progress["maximum"])
        self.run_btn.configure(state="normal")

    # ------------------------------------------------------------------
    # Table helpers
    # ------------------------------------------------------------------
    def _clear_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._defect_rows = []
        self._cat_items = {}

    def _append_row(self, row: dict, row_idx: int, progress_val: int):
        tag = row.get("Disposition", "Needs review")
        cat = category_for_checker(row.get("Checker", ""))
        parent = self._cat_items.get(cat)
        if parent is None or not self.tree.exists(parent):
            # Insert the category header in its canonical CATEGORY_ORDER slot.
            index = sum(1 for c in CATEGORY_ORDER if c in self._cat_items
                        and CATEGORY_ORDER.index(c) < CATEGORY_ORDER.index(cat))
            parent = self.tree.insert("", index, iid=f"cat-{_iid_safe(cat)}",
                                      values=("", cat), tags=("cat_header",))
            self._cat_items[cat] = parent
        self.tree.insert(
            parent, "end", iid=f"r{row_idx}",
            values=(row["CID"], row["Checker"], row["Category"], row["File"],
                    row["Line"], row["Disposition"], row["Confidence"], row["Comment"]),
            tags=(tag,),
        )
        # Keep the live running count on the category header.
        self.tree.item(parent, values=("", f"{cat}  ({len(self.tree.get_children(parent))})"))
        self.progress.configure(value=progress_val)

    def _refresh_row(self, item_id: str, row: dict):
        tag = row.get("Disposition", "Needs review")
        self.tree.item(item_id, values=(
            row["CID"], row["Checker"], category_for_checker(row.get("Checker", "")),
            row["File"], row["Line"], row["Disposition"], row["Confidence"], row["Comment"],
        ), tags=(tag,))

    def _sort_by(self, col: str):
        parents = self.tree.get_children("")
        # Grouped view: sort only the leaf rows inside each category header,
        # leaving header order (CATEGORY_ORDER) untouched.
        if any(self.tree.get_children(p) for p in parents):
            for p in parents:
                data = [(self.tree.set(child, col), child)
                        for child in self.tree.get_children(p)]
                try:
                    data.sort(key=lambda t: int(t[0]) if t[0].isdigit() else t[0].lower())
                except Exception:
                    data.sort(key=lambda t: t[0].lower())
                for index, (_, child) in enumerate(data):
                    self.tree.move(child, p, index)
            return

        data = [(self.tree.set(child, col), child) for child in parents]
        try:
            data.sort(key=lambda t: int(t[0]) if t[0].isdigit() else t[0].lower())
        except Exception:
            data.sort(key=lambda t: t[0].lower())
        for index, (_, child) in enumerate(data):
            self.tree.move(child, "", index)

    # ------------------------------------------------------------------
    # Row double-click — edit disposition & comment
    # ------------------------------------------------------------------
    def _on_row_double_click(self, event):
        item = self.tree.focus()
        if not item:
            return
        if self.tree.parent(item) == "":
            # Category header — toggle expand/collapse.
            self.tree.item(item, open=not self.tree.item(item, "open"))
            return
        # Leaf rows use iid "r<idx>" so we map straight back to self._defect_rows.
        idx = int(item[1:]) if item.startswith("r") and item[1:].isdigit() else self.tree.index(item)
        if idx >= len(self._defect_rows):
            return
        row = self._defect_rows[idx]
        self._open_edit_window(item, idx, row)

    def _open_edit_window(self, item_id: str, idx: int, row: dict):
        win = tk.Toplevel(self.root)
        win.title(f"Defect CID {row['CID']} — Edit")
        win.geometry("720x560")
        win.grab_set()

        # Info header
        ttk.Label(win, text=f"CID {row['CID']}  |  {row['Checker']}  |  {row['File']}:{row['Line']}",
                  font=("Arial", 10, "bold")).pack(anchor="w", padx=10, pady=(10, 2))

        # Events
        ttk.Label(win, text="Coverity Events:", font=("Arial", 9, "underline")).pack(anchor="w", padx=10, pady=(8, 0))
        events_box = scrolledtext.ScrolledText(win, height=5, wrap="word", state="normal")
        events_box.pack(fill="x", padx=10)
        for evt in row.get("_events", []):
            events_box.insert("end",
                f"[{evt.get('tag','?')}] {evt.get('file','')}:{evt.get('line','')}  {evt.get('description','')}\n")
        events_box.configure(state="disabled")

        # Disposition selector
        disp_frame = ttk.Frame(win)
        disp_frame.pack(fill="x", padx=10, pady=(10, 0))
        ttk.Label(disp_frame, text="Disposition:").pack(side="left")
        disp_var = tk.StringVar(value=row["Disposition"])
        disp_combo = ttk.Combobox(disp_frame, textvariable=disp_var,
                                  values=DISPOSITION_OPTIONS, state="readonly", width=20)
        disp_combo.pack(side="left", padx=6)

        # Comment
        ttk.Label(win, text="Comment:").pack(anchor="w", padx=10, pady=(8, 0))
        comment_box = scrolledtext.ScrolledText(win, height=5, wrap="word")
        comment_box.insert("1.0", row["Comment"])
        comment_box.pack(fill="x", padx=10)

        # Source code preview
        ttk.Label(win, text="Source Code Context:", font=("Arial", 9, "underline")).pack(anchor="w", padx=10, pady=(8, 0))
        code_box = scrolledtext.ScrolledText(win, height=8, wrap="none",
                                             font=("Courier New", 9))
        code_box.insert("1.0", self._get_source_snippet(row))
        code_box.configure(state="disabled")
        code_box.pack(fill="both", expand=True, padx=10)

        # Buttons
        btn_frame = ttk.Frame(win)
        btn_frame.pack(fill="x", padx=10, pady=10)

        def _save():
            new_disp = disp_var.get()
            new_comment = comment_box.get("1.0", "end-1c").strip()
            row["Disposition"] = new_disp
            row["Comment"] = new_comment
            self._defect_rows[idx] = row
            self._refresh_row(item_id, row)
            win.destroy()

        ttk.Button(btn_frame, text="Save", command=_save).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Cancel", command=win.destroy).pack(side="left", padx=5)

    def _get_source_snippet(self, row: dict) -> str:
        file_path = row.get("File", "")
        line = row.get("Line", 0)
        src_root = self.src_root_var.get().strip() or self.report_var.get().strip()

        if not file_path:
            return "(no file info)"
        if not os.path.isabs(file_path):
            file_path = os.path.join(src_root, file_path)
        if not os.path.exists(file_path):
            return f"File not found: {file_path}"
        try:
            line = int(line)
            with open(file_path, "r", encoding="utf-8", errors="replace") as fh:
                lines = fh.readlines()
            start = max(0, line - 15)
            end = min(len(lines), line + 15)
            numbered = [
                f"{start + i + 1:>5} {'>>>' if (start + i + 1 == line) else '   '} {l}"
                for i, l in enumerate(lines[start:end])
            ]
            return "".join(numbered)
        except Exception as exc:
            return f"Error reading source: {exc}"

    # ------------------------------------------------------------------
    # Status helper
    # ------------------------------------------------------------------
    def _set_status(self, msg: str):
        self.status_var.set(msg)

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def _export_excel(self):
        if not _HAS_OPENPYXL:
            messagebox.showerror(
                "Missing library",
                "openpyxl is not installed.\n\nRun:  pip install openpyxl"
            )
            return
        if not self._defect_rows:
            messagebox.showwarning("No Data", "Run the analysis first.")
            return

        out_path = filedialog.asksaveasfilename(
            title="Save Excel Report",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile="coverity_dispositions.xlsx",
        )
        if not out_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Coverity Triage"

            # ---- Header row ----
            headers = ["CID", "Checker", "Category", "File", "Line", "Disposition", "Confidence", "Comment",
                       "Events Summary"]
            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align

            ws.row_dimensions[1].height = 22

            # ---- Column widths ----
            excel_col_widths = {1: 8, 2: 22, 3: 18, 4: 45, 5: 7, 6: 18, 7: 12, 8: 50, 9: 60}
            for col_idx, width in excel_col_widths.items():
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = width

            # Effective width for wrapped comment text. We deliberately do NOT
            # fill the full column width (50) and do NOT use a scrollable model;
            # the comment is wrapped to the remaining usable space minus a margin.
            COMMENT_WRAP_WIDTH = 42
            COMMENT_MAX_LINES = 12   # cap so the block never consumes full height

            # ---- Disposition fill colours ----
            disp_fills = {
                "Bug":            PatternFill("solid", fgColor="FFB3B3"),
                "False positive": PatternFill("solid", fgColor="B3FFB3"),
                "Intentional":    PatternFill("solid", fgColor="B3D1FF"),
                "Needs review":   PatternFill("solid", fgColor="FFF3B3"),
            }
            wrap = Alignment(wrap_text=True, vertical="top")

            # ---- Data rows ----

            import textwrap as _twrap
            for row_idx, row in enumerate(self._defect_rows, start=2):
                events_summary = "; ".join(
                    f"{e.get('tag','?')}@{e.get('file','')}:{e.get('line','')}"
                    f" — {e.get('description','')}"
                    for e in row.get("_events", [])
                )
                # Comment (col 7): wrap to the remaining usable space but do NOT
                # fill the full column width — leave a margin (no scrollable/fill
                # model). Cap line count so the cell never consumes full height.
                _comment = row.get("Comment") or ""
                if isinstance(_comment, str) and _comment.strip():
                    _wrapped = _twrap.wrap(_comment, width=COMMENT_WRAP_WIDTH,
                                          break_long_words=True, break_on_hyphens=False)
                    row["_wrapped_comment"] = "\n".join(_wrapped[:COMMENT_MAX_LINES])
                else:
                    row["_wrapped_comment"] = _comment
                values = [
                    row["CID"], row["Checker"],
                    row.get("Category") or category_for_checker(row.get("Checker", "")),
                    row["File"], row["Line"],
                    row["Disposition"], row["Confidence"], row["_wrapped_comment"], events_summary,
                ]
                disp = row.get("Disposition", "Needs review")
                row_fill = disp_fills.get(disp)

                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = wrap
                    if row_fill:
                        cell.fill = row_fill

            # ---- Summary sheet ----
            ws_sum = wb.create_sheet("Summary")
            ws_sum.column_dimensions["A"].width = 22
            ws_sum.column_dimensions["B"].width = 12
            ws_sum.cell(row=1, column=1, value="Disposition").font = Font(bold=True)
            ws_sum.cell(row=1, column=2, value="Count").font = Font(bold=True)

            from collections import Counter
            counts = Counter(r["Disposition"] for r in self._defect_rows)
            for r_idx, (disp, cnt) in enumerate(sorted(counts.items()), start=2):
                ws_sum.cell(row=r_idx, column=1, value=disp)
                ws_sum.cell(row=r_idx, column=2, value=cnt)
                fill = disp_fills.get(disp)
                if fill:
                    ws_sum.cell(row=r_idx, column=1).fill = fill
                    ws_sum.cell(row=r_idx, column=2).fill = fill

            ws_sum.cell(row=r_idx + 2, column=1, value="Total").font = Font(bold=True)
            ws_sum.cell(row=r_idx + 2, column=2, value=sum(counts.values())).font = Font(bold=True)

            # ---- Category summary block (columns D/E) ----
            ws_sum.column_dimensions["D"].width = 30
            ws_sum.column_dimensions["E"].width = 12
            ws_sum.cell(row=1, column=4, value="Category").font = Font(bold=True)
            ws_sum.cell(row=1, column=5, value="Count").font = Font(bold=True)
            cat_counts = Counter(category_for_checker(r.get("Checker", ""))
                                 for r in self._defect_rows)
            for c_idx, (catn, cnt) in enumerate(cat_counts.most_common(), start=2):
                ws_sum.cell(row=c_idx, column=4, value=catn)
                ws_sum.cell(row=c_idx, column=5, value=cnt)
            ws_sum.cell(row=c_idx + 2, column=4, value="Total").font = Font(bold=True)
            ws_sum.cell(row=c_idx + 2, column=5, value=sum(cat_counts.values())).font = Font(bold=True)

            ws_sum.cell(row=r_idx + 3, column=1,
                        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

            wb.save(out_path)
            messagebox.showinfo("Exported", f"Excel report saved:\n{out_path}")
            self._set_status(f"Excel exported → {out_path}")

        except Exception as exc:
            messagebox.showerror("Export Error", str(exc))


# ===========================================================================
def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    root = tk.Tk()
    app = CoverityExcelApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()