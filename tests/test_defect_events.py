"""Defect line + Coverity event-trace resolution (706 vs 710).

Coverity Connect shows the *main event* line (710). SOAP/REST v1 lineNumber is
often the first path event (706, e.g. var_decl). Analysis must use 710.
"""
import json
import os
import tempfile

from coverity_events import (
    apply_events_to_defect,
    events_from_json,
    events_from_summary,
    events_to_json,
    is_sink_event,
    line_from_events,
    mark_main_events,
)
from coverity_rest_client import apply_rest_lines
from coverity_soap_client import _line_from_events, _rest_defect_from_issue
from html_report_parser import parse_coverity_excel, parse_detail_page, write_pull_excel


def _trace():
    """Classic Coverity path: declaration at 706, overrun sink at 710."""
    return [
        {"step": 1, "type": "var_decl", "line": 706, "main": False,
         "description": 'Variable "buf" declared here.', "file": "src/foo.c"},
        {"step": 2, "type": "overrun-local", "line": 710, "main": True,
         "description": 'Overrunning array "buf" of 10 bytes.', "file": "src/foo.c"},
    ]


def test_line_from_events_prefers_main_event_not_first_event():
    assert line_from_events(_trace(), checker="OVERRUN") == 710
    assert _line_from_events(_trace(), checker="OVERRUN") == 710


def test_line_from_events_uses_sink_tag_when_main_flag_missing():
    events = [
        {"step": 1, "type": "var_decl", "line": 706, "main": False},
        {"step": 2, "type": "overrun-local", "line": 710, "main": False},
    ]
    assert line_from_events(events, checker="OVERRUN") == 710


def test_line_from_events_last_event_when_no_main_or_sink():
    events = [
        {"step": 1, "type": "cond_true", "line": 706},
        {"step": 2, "type": "identity_transfer", "line": 710},
    ]
    assert line_from_events(events) == 710


def test_apply_events_overrides_instance_occurrence_line():
    defect = {
        "cid": 42,
        "checker": "OVERRUN",
        "file": "src/foo.c",
        "line": 706,              # SOAP instance/merged first-event line
        "_merged_line": 706,
        "_inst_line_val": 706,
    }
    apply_events_to_defect(defect, _trace())
    assert defect["line"] == 710
    assert defect["_line_src"] == "main_event"
    assert len(defect["events"]) == 2
    assert defect["events"][1]["type"] == "overrun-local"


def test_rest_v1_line_number_does_not_clobber_main_event():
    defects = [{
        "cid": 42, "line": 710, "_line_src": "main_event",
        "checker": "OVERRUN", "file": "src/foo.c",
    }]
    # REST v1 /defects only has occurrence lineNumber = 706
    cid_map = {42: {"line": 706, "mainEventLineNumber": 0}}
    apply_rest_lines(defects, cid_map)
    assert defects[0]["line"] == 710


def test_rest_main_event_line_number_is_used():
    defects = [{"cid": 42, "line": 706, "_line_src": "instance"}]
    cid_map = {42: {"line": 710, "mainEventLineNumber": 710}}
    apply_rest_lines(defects, cid_map)
    assert defects[0]["line"] == 710
    assert defects[0]["_line_src"] == "rest_main"


def test_rest_defect_from_issue_prefers_main_event_line_number():
    issue = {
        "cid": 42,
        "checkerName": "OVERRUN",
        "mainEventFilePath": "src/foo.c",
        "mainEventLineNumber": 710,
        "lineNumber": 706,
        "functionDisplayName": "vulnerable_copy",
        "displayType": "Out-of-bounds access",
        "displayImpact": "High",
        "events": [
            {"eventNumber": 1, "eventTag": "var_decl", "lineNumber": 706,
             "main": False, "eventDescription": "declared here"},
            {"eventNumber": 2, "eventTag": "overrun-local", "lineNumber": 710,
             "main": True, "eventDescription": "overrun here"},
        ],
    }
    d = _rest_defect_from_issue(issue, is_v1=False)
    assert d["line"] == 710
    assert d["_rest_main_line"] == 710
    assert len(d["events"]) == 2
    assert d["events"][1]["main"] is True


def test_rest_v1_issue_without_main_event_keeps_occurrence_until_events():
    issue = {
        "checkerName": "OVERRUN",
        "filePathname": "src/foo.c",
        "lineNumber": 706,
    }
    d = _rest_defect_from_issue(issue, is_v1=True)
    assert d["line"] == 706
    assert d["events"] == []


def test_excel_roundtrip_preserves_events_and_main_line(tmp_path):
    defects = [{
        "cid": 42,
        "checker": "OVERRUN",
        "type": "Out-of-bounds access",
        "severity": "High",
        "file": "src/foo.c",
        "line": 710,
        "function": "vulnerable_copy",
        "events": _trace(),
    }]
    out = tmp_path / "pull.xlsx"
    write_pull_excel(defects, str(out))
    parsed = parse_coverity_excel(str(out))
    assert len(parsed) == 1
    d = parsed[0]
    assert d["cid"] == 42
    assert d["line"] == 710
    assert len(d["events"]) == 2
    assert d["events"][0]["line"] == 706
    assert d["events"][1]["line"] == 710
    assert d["events"][1]["main"] is True


def test_events_json_roundtrip():
    blob = events_to_json(_trace())
    back = events_from_json(blob)
    assert line_from_events(back) == 710
    assert json.loads(blob)[1]["line"] == 710


def test_events_from_summary_legacy_column():
    text = (
        'var_decl@src/foo.c:706 — Variable "buf" declared here.; '
        'overrun-local@src/foo.c:710 — Overrunning array "buf"'
    )
    events = events_from_summary(text)
    assert line_from_events(events, "OVERRUN") == 710


def test_html_detail_parser_attaches_event_lines_and_marks_main():
    html = """<html><body><pre>
(1) Event var_decl:
Variable "buf" declared here.
  706      char buf[10];
(2) Event overrun-local:
Overrunning array "buf" of 10 bytes by passing it to memcpy.
  710      memcpy(buf, src, n);
</pre></body></html>"""
    with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False,
                                     encoding="utf-8") as fh:
        fh.write(html)
        path = fh.name
    try:
        code, events = parse_detail_page(path)
        assert len(events) == 2
        assert events[0]["type"] == "var_decl"
        assert events[0]["line"] == 706
        assert events[1]["type"] == "overrun-local"
        assert events[1]["line"] == 710
        assert events[1]["main"] is True
        assert "memcpy" in events[1]["description"]
        assert "710" in code
    finally:
        os.unlink(path)


def test_html_same_line_description_and_hyphenated_tag():
    html = """<html><body><pre>
(1) Event overrun-buffer-arg: Calling 'strncpy' with a maximum size argument of 64 bytes on destination array 'buf' of size 64 bytes.
  11  strncpy(buf, input, sizeof(buf));
</pre></body></html>"""
    with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False,
                                     encoding="utf-8") as fh:
        fh.write(html)
        path = fh.name
    try:
        _code, events = parse_detail_page(path)
        assert len(events) == 1
        assert events[0]["type"] == "overrun-buffer-arg"
        assert events[0]["line"] == 11
        assert "strncpy" in events[0]["description"]
        assert events[0]["main"] is True
    finally:
        os.unlink(path)


def test_string_null_source_is_not_the_sink():
    """STRING_NULL's trace is var_decl -> string_null_source -> string_null.

    ``string_null_source`` (where the string loses its NUL terminator) is a
    *source* event, not the sink. It shares the checker prefix and, as a child
    event, can carry a higher eventNumber than the real ``string_null`` sink —
    so a naive prefix match reports the source line (174) instead of the sink
    line (268).
    """
    assert is_sink_event({"type": "string_null_source"}, "STRING_NULL") is False
    assert is_sink_event({"type": "string_null"}, "STRING_NULL") is True


def test_line_from_events_prefers_string_null_sink_over_source():
    # Source event discovered AFTER the sink (higher eventNumber) — exactly the
    # interprocedural child-event ordering that used to yield the wrong line.
    events = [
        {"step": 1, "type": "var_decl", "line": 174, "main": False},
        {"step": 2, "type": "string_null", "line": 268, "main": False,
         "description": "Passing unterminated string to strlen"},
        {"step": 3, "type": "string_null_source", "line": 174, "main": False,
         "description": "strncpy does not terminate the string"},
    ]
    assert line_from_events(events, checker="STRING_NULL") == 268


def test_apply_events_marks_string_null_sink_main_not_source():
    defect = {
        "cid": 7,
        "checker": "STRING_NULL",
        "file": "src/foo.c",
        "line": 268,
    }
    events = [
        {"step": 1, "type": "var_decl", "line": 174, "main": False},
        {"step": 2, "type": "string_null", "line": 268, "main": False},
        {"step": 3, "type": "string_null_source", "line": 174, "main": False},
    ]
    apply_events_to_defect(defect, events)
    assert defect["line"] == 268
    assert defect["_line_src"] == "main_event"
    main_flags = {e["type"]: e.get("main") for e in defect["events"]}
    assert main_flags["string_null"] is True
    assert main_flags["string_null_source"] is False
    assert main_flags["var_decl"] is False


def test_line_from_events_prefers_buffer_size_over_warning():
    """BUFFER_SIZE: buffer_size (268) must beat buffer_size_warning (174) even
    when the warning has a higher eventNumber (it sits at the declaration)."""
    events = [
        {"step": 1, "type": "var_decl", "line": 174, "main": False},
        {"step": 2, "type": "buffer_size", "line": 268, "main": False,
         "description": "Calling 'strncpy' with size argument..."},
        {"step": 3, "type": "buffer_size_warning", "line": 174, "main": False,
         "description": "warning"},
    ]
    assert line_from_events(events, checker="BUFFER_SIZE") == 268


def test_apply_events_marks_buffer_size_main_not_warning():
    defect = {
        "cid": 8,
        "checker": "BUFFER_SIZE",
        "file": "src/foo.c",
        "line": 268,
    }
    events = [
        {"step": 1, "type": "var_decl", "line": 174, "main": False},
        {"step": 2, "type": "buffer_size", "line": 268, "main": False},
        {"step": 3, "type": "buffer_size_warning", "line": 174, "main": False},
    ]
    apply_events_to_defect(defect, events)
    assert defect["line"] == 268
    main_flags = {e["type"]: e.get("main") for e in defect["events"]}
    assert main_flags["buffer_size"] is True
    assert main_flags["buffer_size_warning"] is False


def test_overrun_child_sink_does_not_shadow_main():
    """OVERRUN: a second sink-tagged event in a callee (higher eventNumber)
    must not shadow the primary overrun-local sink at 268."""
    events = [
        {"step": 1, "type": "var_decl", "line": 174, "main": False},
        {"step": 2, "type": "overrun-local", "line": 268, "main": False},
        {"step": 3, "type": "overrun-buffer-arg", "line": 500, "main": False},
    ]
    assert line_from_events(events, checker="OVERRUN") == 268


def test_warning_and_source_are_not_primary_sinks():
    assert is_sink_event({"type": "buffer_size_warning"}, "BUFFER_SIZE") is False
    assert is_sink_event({"type": "buffer_size"}, "BUFFER_SIZE") is True
    # Exact checker tag wins even for the warning checker itself.
    assert is_sink_event({"type": "buffer_size_warning"}, "BUFFER_SIZE_WARNING") is True
    assert is_sink_event({"type": "string_null_source"}, "STRING_NULL") is False


def test_heuristic_main_is_recomputed_when_checker_known():
    """A checker-less marking (e.g. from HTML parsing) that picked the wrong
    event must be corrected once the checker is known."""
    events = [
        {"step": 1, "type": "var_decl", "line": 174, "main": False},
        {"step": 2, "type": "buffer_size", "line": 268, "main": False},
        {"step": 3, "type": "buffer_size_warning", "line": 174, "main": False},
    ]
    # First marking pass without a checker (simulates parse_detail_page).
    mark_main_events(events)
    # Then the checker-aware pass must re-evaluate, not trust the old flag.
    mark_main_events(events, checker="BUFFER_SIZE")
    main_flags = {e["type"]: e.get("main") for e in events}
    assert main_flags["buffer_size"] is True
    assert main_flags["buffer_size_warning"] is False


def test_genuine_main_flag_is_preserved():
    """Coverity's own main=true must not be overridden by a re-marking pass."""
    events = [
        {"step": 1, "type": "var_decl", "line": 174, "main": False},
        {"step": 2, "type": "buffer_size_warning", "line": 174, "main": True},
        {"step": 3, "type": "buffer_size", "line": 268, "main": False},
    ]
    mark_main_events(events, checker="BUFFER_SIZE")
    main_flags = {e["type"]: e.get("main") for e in events}
    assert main_flags["buffer_size_warning"] is True
    assert main_flags["buffer_size"] is False
    assert line_from_events(events, checker="BUFFER_SIZE") == 174


def test_mark_main_events_picks_sink():
    events = [
        {"step": 1, "type": "var_decl", "line": 706, "main": False},
        {"step": 2, "type": "overrun-local", "line": 710, "main": False},
    ]
    mark_main_events(events, checker="OVERRUN")
    assert events[1]["main"] is True
    assert events[0]["main"] is False


def test_pull_excel_writes_eventsjson_column(tmp_path):
    """write_pull_excel must actually serialize the event trace into the
    EventsJSON column (it used to declare the column but never write it,
    so the main-event line was lost in the Excel round-trip)."""
    import openpyxl
    defects = [{
        "cid": 42, "checker": "OVERRUN", "type": "Out-of-bounds access",
        "severity": "High", "file": "src/foo.c", "line": 710,
        "function": "vulnerable_copy", "events": _trace(),
    }]
    out = tmp_path / "pull.xlsx"
    write_pull_excel(defects, str(out))
    wb = openpyxl.load_workbook(str(out), data_only=True)
    ws = wb.active
    headers = [str(c.value) for c in ws[1]]
    idx = headers.index("EventsJSON")
    blob = ws.cell(row=2, column=idx + 1).value
    assert blob
    events = events_from_json(blob)
    assert len(events) == 2
    assert events[1]["line"] == 710
    assert events[1]["main"] is True

