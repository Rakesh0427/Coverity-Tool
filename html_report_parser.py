"""
Coverity HTML Report Parser - two-phase API for live progress updates.
Phase 1: parse_index_only()   - reads only index.html (fast)
Phase 2: parse_detail_page()  - reads one Code/*.html file per defect
Auto-fallback: lxml -> html.parser if lxml is not installed.
Recursively searches all subdirectories for detail HTML files.
"""
import os, re
from bs4 import BeautifulSoup
from typing import List, Dict, Any, Tuple

_PARSER = "html.parser"
try:
    import lxml
    _PARSER = "lxml"
except ImportError:
    pass

# Cache for discovered detail files across subfolders
_DETAIL_FILE_CACHE = {}


def _soup(path: str) -> BeautifulSoup:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        return BeautifulSoup(f, _PARSER)


def _build_detail_file_cache(base_dir: str):
    """Recursively scan base_dir and subdirectories to map .html filenames to full paths."""
    global _DETAIL_FILE_CACHE
    _DETAIL_FILE_CACHE = {}
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.endswith(".html"):
                full = os.path.join(root, f)
                _DETAIL_FILE_CACHE[f] = full
                rel = os.path.relpath(full, base_dir).replace("\\", "/")
                _DETAIL_FILE_CACHE[rel] = full


def _extract_line_from_file(text):
    """Extract line number from combined file+line text like '/path/file.cpp 123'."""
    if not text:
        return text, 0
    m = re.search(r'^(.*?)\s*[:\s]\s*(\d+)\s*$', text)
    if m:
        return m.group(1).strip(), int(m.group(2))
    return text, 0


def _is_checker(text):
    return bool(re.match(r'^[A-Z][A-Z_\d]+$', text)) and len(text) > 3


def _is_file_path(text):
    return ('/' in text or '\\' in text) and len(text) > 3


def _is_line_number(text):
    return bool(re.match(r'^\d{1,5}$', text)) and int(text) < 100000


def _is_severity(text):
    return text.lower() in ('high', 'medium', 'low', 'critical', 'info', 'unspecified')


def _is_cid(text):
    return bool(re.match(r'^\d+$', text))


def _is_hex_hash(text):
    """Reject 32+ char hex strings (MD5/SHA1 hashes) and UUID-like strings."""
    if len(text) >= 16 and re.match(r'^[0-9a-fA-F]+$', text):
        return True
    return False


def _is_non_function_word(text):
    """Reject literal status words that Coverity puts in function column."""
    return text.lower() in ('unavailable', 'n/a', 'na', 'none', 'unknown', 'unclassified')


def parse_index_only(report_path: str) -> List[Dict[str, Any]]:
    """
    Read only index.html and return lightweight defect dicts.
    Columns detected by content patterns (works with any Coverity layout).
    Each dict has: cid, checker, type, severity, file, line,
                   function, detail_file (absolute path or '' if missing).
    """
    if os.path.isdir(report_path):
        report_path = os.path.join(report_path, "index.html")
    if not os.path.isfile(report_path):
        raise FileNotFoundError(f"Report not found: {report_path}")

    base_dir = os.path.dirname(os.path.abspath(report_path))

    # Build recursive cache of all .html files (handles subfolders like 1/, 2/, Code/1/ etc.)
    _build_detail_file_cache(base_dir)

    soup = _soup(report_path)

    table = soup.find("table")
    if not table:
        raise ValueError("No <table> found in index.html")

    rows = table.find_all("tr")
    if not rows:
        return []

    # ---------- Detect "Type" and "Function" columns from header ----------
    type_col = None
    func_col = None
    header_cells = []
    if rows:
        first = rows[0]
        ths = first.find_all("th")
        if ths:
            header_cells = [th.get_text(strip=True).lower() for th in ths]
        else:
            tds = first.find_all("td")
            if tds:
                first_text = tds[0].get_text(strip=True).lower()
                if first_text in ("cid", "id", "defect id", "checker", "type", "#"):
                    header_cells = [td.get_text(strip=True).lower() for td in tds]
        if header_cells:
            for idx, h in enumerate(header_cells):
                if h == "type":
                    type_col = idx
                if h in ("function", "func", "function name", "function_name"):
                    func_col = idx

    # Skip header row if present
    start_idx = 0
    if rows[0].find_all("th"):
        start_idx = 1
    else:
        first_cells = rows[0].find_all("td")
        if first_cells:
            first_text = first_cells[0].get_text(strip=True).lower()
            if first_text in ("cid", "id", "defect id", "checker", "type", "#"):
                start_idx = 1

    defects = []

    for row in rows[start_idx:]:
        cells = row.find_all("td")
        if len(cells) < 3:
            continue

        cell_texts = [c.get_text(strip=True) for c in cells]

        # --- CID: first pure integer cell ---
        cid = None
        cid_idx = -1
        for i, text in enumerate(cell_texts):
            if _is_cid(text):
                cid = int(text)
                cid_idx = i
                break
        if cid is None:
            continue

        # --- Checker: first UPPER_CASE cell after CID ---
        checker = ""
        for i, text in enumerate(cell_texts):
            if i == cid_idx:
                continue
            if _is_checker(text):
                checker = text
                break

        # --- File: first cell with path separator ---
        file_path = ""
        line = 0
        file_idx = -1
        for i, text in enumerate(cell_texts):
            if i == cid_idx:
                continue
            if _is_file_path(text):
                file_path, extracted_line = _extract_line_from_file(text)
                if extracted_line > 0:
                    line = extracted_line
                file_idx = i
                break

        # --- Line: if not extracted from file, look for standalone line number ---
        if line == 0:
            for i, text in enumerate(cell_texts):
                if i in (cid_idx, file_idx):
                    continue
                if _is_line_number(text):
                    line = int(text)
                    break

        # --- Severity: known values ---
        severity = ""
        for text in cell_texts:
            if _is_severity(text):
                severity = text
                break

        # --- Function: use header column if available, else heuristic ---
        function = ""
        if func_col is not None and func_col < len(cells):
            raw_func = cells[func_col].get_text(strip=True)
            clean = re.sub(r'\(.*\)\s*$', '', raw_func)
            if (clean and not _is_non_function_word(clean) and not _is_hex_hash(clean) and
                re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', clean) and len(clean) < 60):
                function = clean

        # Fallback heuristic only if header didn't give us a valid function
        if not function:
            for text in cell_texts:
                clean = re.sub(r'\(.*\)\s*$', '', text.strip())
                if (clean and not _is_non_function_word(clean) and not _is_hex_hash(clean) and
                    clean != checker and clean != file_path and clean != severity and
                    not _is_cid(clean) and not _is_line_number(clean) and
                    not _is_file_path(clean) and not _is_checker(clean) and
                    re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', clean) and 2 < len(clean) < 60):
                    function = clean
                    break

        # --- Detail link ---
        link = cells[cid_idx].find("a") if cid_idx < len(cells) else None
        if not link:
            link = row.find("a")
        href = link.get("href", "") if link else ""
        detail_file = ""
        if href:
            p = os.path.normpath(os.path.join(base_dir, href.split("#")[0]))
            if os.path.isfile(p):
                detail_file = p
            else:
                fname = os.path.basename(href.split("#")[0])
                if fname in _DETAIL_FILE_CACHE:
                    detail_file = _DETAIL_FILE_CACHE[fname]
                else:
                    rel = href.split("#")[0].replace("\\", "/")
                    if rel in _DETAIL_FILE_CACHE:
                        detail_file = _DETAIL_FILE_CACHE[rel]
                    else:
                        for key, val in _DETAIL_FILE_CACHE.items():
                            if key.endswith(fname):
                                detail_file = val
                                break

        # --- Type: use header-identified "Type" column, or empty ---
        defect_type = ""
        if type_col is not None and type_col < len(cells):
            defect_type = cells[type_col].get_text(strip=True)

        defects.append({
            "cid": cid, "checker": checker, "type": defect_type,
            "severity": severity, "file": file_path, "line": line,
            "function": function, "detail_file": detail_file,
            "events": [], "source_code": "",
        })
    return defects


def parse_detail_page(detail_path: str) -> Tuple[str, List[Dict]]:
    """
    Parse one Code/<n>_file.html and return (source_code, events).
    Uses the ORIGINAL text‑based strategy that was proven to work.
    """
    if not detail_path or not os.path.isfile(detail_path):
        return "", []
    try:
        soup = _soup(detail_path)
        text_lines = soup.get_text().splitlines()
        source_lines = []
        events = []

        ev_pat = re.compile(r"^\((\d+)\)\s+Event\s+(\w+):$")
        # Flexible code-line pattern: line number (1-6 digits), then whitespace, then code
        code_pat = re.compile(r"^(\d{1,6})\s+(.*)$")

        i = 0
        while i < len(text_lines):
            raw = text_lines[i].strip()
            # -- Event line --
            m = ev_pat.match(raw)
            if m:
                desc = text_lines[i + 1].strip() if i + 1 < len(text_lines) else ""
                events.append({
                    "step": int(m.group(1)),
                    "type": m.group(2),
                    "description": desc
                })
                i += 2
                continue
            # -- Code line --
            cm = code_pat.match(raw)
            if cm:
                source_lines.append(f"{cm.group(1):>6}  {cm.group(2)}")
            i += 1

        return "\n".join(source_lines), events
    except Exception as e:
        return "", [{"step": 1, "type": "parse_error", "description": str(e)}]
    


# ---------------------------------------------------------------------------
# Coverity Excel (Black Duck export) parser
# ---------------------------------------------------------------------------

def _fuzzy_col_match(headers: List[str], candidates: List[str]) -> int:
    """Return column index whose header contains or exactly matches any candidate word."""
    headers_lower = [h.lower().replace('_', ' ').replace('-', ' ') for h in headers]
    # First: exact match
    for cand in candidates:
        cand_clean = cand.lower().strip()
        for idx, h in enumerate(headers_lower):
            if h == cand_clean:
                return idx
    # Second: substring match (candidate contained in header)
    for cand in candidates:
        cand_clean = cand.lower().strip()
        for idx, h in enumerate(headers_lower):
            if cand_clean in h:
                return idx
    # Third: header contained in candidate (reverse)
    for cand in candidates:
        cand_clean = cand.lower().strip()
        for idx, h in enumerate(headers_lower):
            if h in cand_clean and len(h) > 1:
                return idx
    return -1


def parse_coverity_excel(excel_path: str) -> List[Dict[str, Any]]:
    """
    Parse a Coverity / Black Duck exported Excel file.
    Handles 'Various' line numbers by storing line=0 and flagging it.
    Returns same dict format as parse_index_only for compatibility.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel support. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h) if h is not None else "" for h in rows[0]]

    # Fuzzy column mapping
    col_cid       = _fuzzy_col_match(headers, ["cid", "defect id", "issue id", "defect", "id"])
    col_checker   = _fuzzy_col_match(headers, ["checker", "type", "issue type", "checker type"])
    col_type      = _fuzzy_col_match(headers, ["type", "subtype", "checker type"])
    col_severity  = _fuzzy_col_match(headers, ["severity", "impact", "priority"])
    col_file      = _fuzzy_col_match(headers, ["file", "source file", "path", "filepath", "file path"])
    col_line      = _fuzzy_col_match(headers, ["line", "line number", "location", "lineno"])
    col_function  = _fuzzy_col_match(headers, ["function", "function name", "procedure", "func"])

    if col_cid == -1:
        raise ValueError(f"Could not find CID/Defect ID column in Excel. Headers: {headers}")
    if col_file == -1:
        raise ValueError(f"Could not find File column in Excel. Headers: {headers}")

    defects = []
    for row in rows[1:]:
        if not row or len(row) <= max(col_cid, col_file):
            continue

        cid_val = row[col_cid]
        try:
            cid = int(cid_val)
        except (ValueError, TypeError):
            continue

        file_val = str(row[col_file]) if row[col_file] is not None else ""
        line_val = row[col_line] if col_line != -1 and col_line < len(row) else ""
        func_val = str(row[col_function]) if col_function != -1 and col_function < len(row) and row[col_function] is not None else ""

        # Handle "Various" line numbers
        line_is_various = False
        line = 0
        if line_val is not None:
            line_str = str(line_val).strip()
            if line_str.lower() == "various":
                line_is_various = True
                line = 0
            else:
                try:
                    line = int(line_str)
                except (ValueError, TypeError):
                    line = 0

        # Clean up function name
        if func_val:
            func_val = re.sub(r'\(.*\)\s*$', '', func_val.strip())
            if func_val.lower() in ('unclassified', 'unavailable', 'none', 'unknown', 'na', 'n/a'):
                func_val = ""

        checker_val = str(row[col_checker]) if col_checker != -1 and col_checker < len(row) and row[col_checker] is not None else ""
        type_val = str(row[col_type]) if col_type != -1 and col_type < len(row) and row[col_type] is not None else ""
        severity_val = str(row[col_severity]) if col_severity != -1 and col_severity < len(row) and row[col_severity] is not None else ""

        defects.append({
            "cid": cid,
            "checker": checker_val,
            "type": type_val,
            "severity": severity_val,
            "file": file_val,
            "line": line,
            "line_is_various": line_is_various,
            "function": func_val,
            "detail_file": "",
            "events": [],
            "source_code": "",
        })

    return defects

def write_pull_excel(defects: List[Dict[str, Any]], output_path: str) -> None:
    """
    Write a structured .xlsx file from SOAP-pulled defects.

    Columns are named to match the fuzzy column matcher in parse_coverity_excel()
    so no parser changes are required. The 'Line' column is always an integer
    (never "Various") — the core fix of the pull-from-Coverity feature.

    Raises on failure (caller catches and logs it).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Coverity"

    headers = ["CID", "Checker", "Subtype", "Severity", "File",
               "Line", "Function", "Events Summary"]
    widths  = [8, 22, 25, 12, 45, 7, 30, 60]

    # ---- Header row ----
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, (hdr, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill    = hdr_fill
        cell.font    = hdr_font
        cell.alignment = hdr_align
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width
    ws.row_dimensions[1].height = 22

    # ---- Data rows ----
    wrap_all = Alignment(horizontal="center", vertical="top")
    wrap_events = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for row_idx, d in enumerate(defects, start=2):
        events_summary = "; ".join(
            f"{e.get('type') or e.get('tag') or '?'}@"
            f"{e.get('file', '')}:{e.get('line', '')}"
            f" — {e.get('description', '')}"
            for e in d.get("events", [])
        )
        values = [
            d.get("cid", ""),
            d.get("checker", ""),
            d.get("type", ""),
            d.get("severity", ""),
            d.get("file", ""),
            d.get("line", 0),
            d.get("function", ""),
            events_summary,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 8:   # Events Summary
                cell.alignment = wrap_events
            else:
                cell.alignment = wrap_all

    wb.save(output_path)

def parse_coverity_html(report_path: str) -> List[Dict[str, Any]]:
    defects = parse_index_only(report_path)
    for d in defects:
        code, events = parse_detail_page(d["detail_file"])
        d["source_code"] = code
        d["events"]      = events or [{"step": 1, "type": d["checker"],
                                        "description": d["type"],
                                        "file": d["file"], "line": d["line"]}]
    return defects  